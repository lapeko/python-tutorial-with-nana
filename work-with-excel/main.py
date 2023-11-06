import openpyxl

inventory_book = openpyxl.load_workbook("inventory.xlsx")

sheet = inventory_book["Sheet1"]

supplier_products = {}
supplier_total = {}
products_under_10 = {}

for row in range(2, sheet.max_row + 1):
    product_cell = sheet.cell(row, column=1)
    inventory_cell = sheet.cell(row, column=2)
    price_cell = sheet.cell(row, column=3)
    supplier_cell = sheet.cell(row, column=4)
    total_value_cell = sheet.cell(row, column=5)

    product = product_cell.value
    inventory = inventory_cell.value
    price = price_cell.value
    supplier = supplier_cell.value

    supplier_products[supplier] = supplier_products.get(supplier, 0) + 1
    supplier_total[supplier] = supplier_total.get(supplier, inventory * price) + inventory * price
    total_value_cell.value = inventory * price

    if float(inventory) < 10:
        products_under_10[product] = inventory


print(supplier_products)
print(supplier_total)
print(products_under_10)

inventory_book.save("inventory-result.xlsx")
